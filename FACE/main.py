import cv2
import os
import time
import pickle
import threading
import traceback
import hashlib
import mediapipe as mp
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
from ultralytics import YOLO
from deepface import DeepFace
from openpyxl import Workbook, load_workbook

# ============================================================
# CẤU HÌNH HỆ THỐNG
# ============================================================
CAMERA_INDEX = 0
YOLO_MODEL_PATH = "yolov8n-face.pt"
KNOWN_FACES_DIR = "known_faces"
CACHE_FILE = "face_cache.pkl"
EVIDENCE_DIR = "attendance_evidence"
EXCEL_FILENAME = f"DanhSach_DiemDanh_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

WINDOW_NAME = "He Thong Diem Danh Tu Dong"
SHOW_DEBUG = True
USE_ENTRANCE_ROI = True

# Camera / Detector
CAMERA_WIDTH = 1280
CAMERA_HEIGHT = 720
YOLO_CONFIDENCE = 0.45
MIN_FACE_SIZE = 36
ROI_DETECT_ZOOM = 1.6

# Recognition
EMBED_MODEL_NAME = "ArcFace"
FACE_PADDING_RATIO = 0.28
EVIDENCE_PADDING_RATIO = 0.20
RECOGNIZE_EVERY_N_FRAMES = 2
MATCH_THRESHOLD_NEAR = 0.63
MATCH_THRESHOLD_FAR = 0.51
GAP_THRESHOLD_NEAR = 0.05
GAP_THRESHOLD_FAR = 0.08
UNKNOWN_MIN_SIM = 0.34

# Tracking / Confirmation
TRACK_MAX_MISSING = 10
TRACK_MATCH_IOU = 0.30
TRACK_MAX_CENTER_DIST_RATIO = 0.18
REQUIRED_CONFIRMATIONS = 5
ALLOW_ONE_WEAK_FRAME = True

INVALID_NAMES = {"Unknown", "Qua xa", "Loi AI", "Khong chac"}

# ============================================================
# BIẾN TOÀN CỤC VÀ MODEL
# ============================================================
latest_frame = None
latest_display = []
is_running = True
frame_lock = threading.Lock()
result_lock = threading.Lock()

ai_processing = False
ai_fps = 0.0

logged_names = set()
tracks: Dict[int, "TrackState"] = {}
next_track_id = 0

# Khởi tạo MediaPipe FaceMesh (Dùng cho Căn chỉnh khuôn mặt)
mp_face_mesh = mp.solutions.face_mesh
# Dùng static_image_mode vì ta sẽ truyền các ảnh crop rời rạc vào
face_mesh = mp_face_mesh.FaceMesh(
    static_image_mode=True, 
    max_num_faces=1, 
    refine_landmarks=False, 
    min_detection_confidence=0.5
)

# DB khuôn mặt: {name: [embedding1, embedding2, ...]}
known_face_db: Dict[str, List[np.ndarray]] = {}
model = None


# ============================================================
# DATA CLASSES
# ============================================================
@dataclass
class Detection:
    box: Tuple[int, int, int, int]
    confidence: float = 0.0
    name: str = "Unknown"
    score: float = 0.0
    need_recognition: bool = True


@dataclass
class TrackState:
    box: Tuple[int, int, int, int]
    missing: int = 0
    best_name: str = "Unknown"
    best_score: float = 0.0
    confirmed: bool = False
    attendance_marked: bool = False

    streak_name: Optional[str] = None
    streak_count: int = 0
    score_history: List[float] = field(default_factory=list)

    snapshot: Optional[np.ndarray] = None
    snapshot_quality: float = -1.0
    last_update_time: float = field(default_factory=time.time)
    last_recognized_frame: int = -99999


# ============================================================
# MODULE NÂNG CAO CHẤT LƯỢNG ẢNH (ALIGN & GAMMA)
# ============================================================
def auto_gamma_correction(img: np.ndarray) -> np.ndarray:
    """
    Tự động tính toán độ sáng của ảnh và áp dụng Gamma Correction.
    Cứu sáng các khuôn mặt bị tối/ngược sáng một cách tự nhiên.
    """
    if img is None or img.size == 0:
        return img
        
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    mean = np.mean(gray)
    
    # Giới hạn an toàn để tránh chia cho 0
    if mean == 0: mean = 1
    if mean == 255: mean = 254
    
    # Mục tiêu kéo giá trị trung bình về 128 (độ sáng lý tưởng)
    # Phương trình: (mean / 255) ^ gamma = 0.5
    gamma = np.log(0.5) / np.log(mean / 255.0)
    
    # Giới hạn Gamma để ảnh không bị bợt màu quá mức
    gamma = np.clip(gamma, 0.6, 1.8)
    
    invGamma = 1.0 / gamma
    table = np.array([((i / 255.0) ** invGamma) * 255 for i in np.arange(0, 256)]).astype("uint8")
    
    return cv2.LUT(img, table)

def align_face(face_crop: np.ndarray) -> np.ndarray:
    """
    Tìm 2 mắt bằng MediaPipe và xoay thẳng khuôn mặt.
    Khắc phục triệt để lỗi do sinh viên nghiêng đầu.
    """
    if face_crop is None or face_crop.size == 0:
        return face_crop

    rgb_crop = cv2.cvtColor(face_crop, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(rgb_crop)

    # Nếu không tìm thấy các điểm trên mặt, trả về ảnh gốc
    if not results.multi_face_landmarks:
        return face_crop

    h, w = face_crop.shape[:2]
    landmarks = results.multi_face_landmarks[0].landmark

    # Lấy tọa độ 2 đuôi mắt (Điểm 33 và 263 theo tiêu chuẩn MediaPipe)
    left_eye_x, left_eye_y = int(landmarks[33].x * w), int(landmarks[33].y * h)
    right_eye_x, right_eye_y = int(landmarks[263].x * w), int(landmarks[263].y * h)

    # Tính toán góc lệch
    dy = right_eye_y - left_eye_y
    dx = right_eye_x - left_eye_x
    angle = np.degrees(np.arctan2(dy, dx))

    # Tọa độ tâm giữa 2 mắt
    eyes_center = ((left_eye_x + right_eye_x) // 2, (left_eye_y + right_eye_y) // 2)

    # Xoay ảnh để làm thẳng 2 mắt (Dùng BORDER_REPLICATE để bù đắp viền bị khuyết)
    M = cv2.getRotationMatrix2D(eyes_center, angle, 1.0)
    aligned_crop = cv2.warpAffine(face_crop, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)

    return aligned_crop


# ============================================================
# HÀM PHỤ CHUNG
# ============================================================
def ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path)

def l2_normalize(vec: np.ndarray) -> np.ndarray:
    vec = np.asarray(vec, dtype=np.float32)
    norm = np.linalg.norm(vec)
    if norm < 1e-8: return vec
    return vec / norm

def compute_cosine_similarity(feat1: np.ndarray, feat2: np.ndarray) -> float:
    feat1, feat2 = l2_normalize(feat1), l2_normalize(feat2)
    n1, n2 = np.linalg.norm(feat1), np.linalg.norm(feat2)
    if n1 < 1e-8 or n2 < 1e-8: return -1.0
    return float(np.dot(feat1, feat2))

def compute_iou(box_a, box_b) -> float:
    ax1, ay1, ax2, ay2 = box_a
    bx1, by1, bx2, by2 = box_b
    inter_x1 = max(ax1, bx1)
    inter_y1 = max(ay1, by1)
    inter_x2 = min(ax2, bx2)
    inter_y2 = min(ay2, by2)
    inter_w = max(0, inter_x2 - inter_x1)
    inter_h = max(0, inter_y2 - inter_y1)
    inter_area = inter_w * inter_h
    area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
    area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
    union = area_a + area_b - inter_area
    if union <= 0: return 0.0
    return float(inter_area / union)

def box_center(box) -> Tuple[float, float]:
    x1, y1, x2, y2 = box
    return ((x1 + x2) / 2.0, (y1 + y2) / 2.0)

def box_area(box) -> float:
    x1, y1, x2, y2 = box
    return float(max(0, x2 - x1) * max(0, y2 - y1))

def normalized_center_distance(box_a, box_b, frame_w: int, frame_h: int) -> float:
    ax, ay = box_center(box_a)
    bx, by = box_center(box_b)
    dist = ((ax - bx) ** 2 + (ay - by) ** 2) ** 0.5
    diag = max(1.0, (frame_w ** 2 + frame_h ** 2) ** 0.5)
    return float(dist / diag)

def point_in_box(cx, cy, box) -> bool:
    x1, y1, x2, y2 = box
    return x1 <= cx <= x2 and y1 <= cy <= y2

def make_label(name: str, score: float) -> str:
    return f"{name} ({score * 100:.0f}%)" if score > 0 else name

def get_face_threshold(face_w: int) -> float:
    if face_w < 40: return MATCH_THRESHOLD_FAR
    if face_w > 80: return MATCH_THRESHOLD_NEAR
    return MATCH_THRESHOLD_FAR + (MATCH_THRESHOLD_NEAR - MATCH_THRESHOLD_FAR) * ((face_w - 40) / 40.0)

def get_gap_threshold(face_w: int) -> float:
    if face_w < 45: return GAP_THRESHOLD_FAR
    if face_w > 85: return GAP_THRESHOLD_NEAR
    return GAP_THRESHOLD_FAR + (GAP_THRESHOLD_NEAR - GAP_THRESHOLD_FAR) * ((face_w - 45) / 40.0)

def preprocess_for_detection(img: np.ndarray) -> np.ndarray:
    # Sử dụng Auto Gamma cho việc detect YOLO nét hơn trong bóng tối
    return auto_gamma_correction(img)

def get_entrance_roi(frame_w: int, frame_h: int):
    roi_w = int(frame_w * 0.60)
    roi_h = int(frame_h * 0.78)
    x1 = (frame_w - roi_w) // 2
    y1 = int(frame_h * 0.15)
    x2 = x1 + roi_w
    y2 = y1 + roi_h
    return (x1, y1, x2, y2)

def crop_with_padding(frame: np.ndarray, box, pad_ratio: float) -> np.ndarray:
    x1, y1, x2, y2 = box
    h, w = frame.shape[:2]
    bw, bh = x2 - x1, y2 - y1
    pad_w, pad_h = int(bw * pad_ratio), int(bh * pad_ratio)
    x1p, y1p = max(0, x1 - pad_w), max(0, y1 - pad_h)
    x2p, y2p = min(w, x2 + pad_w), min(h, y2 + pad_h)
    crop = frame[y1p:y2p, x1p:x2p]
    return crop.copy() if crop.size > 0 else np.empty((0, 0, 3), dtype=np.uint8)

def compute_blur_score(img: np.ndarray) -> float:
    if img is None or img.size == 0: return 0.0
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return float(cv2.Laplacian(gray, cv2.CV_64F).var())

def compute_face_quality(face_crop: np.ndarray, recog_score: float) -> float:
    if face_crop is None or face_crop.size == 0: return -1.0
    h, w = face_crop.shape[:2]
    area_score = min((h * w) / 15000.0, 2.0)
    blur_score = min(compute_blur_score(face_crop) / 150.0, 3.0)
    recog_bonus = max(0.0, recog_score) * 4.0
    return area_score + blur_score + recog_bonus

def get_dataset_signature(items: List[Tuple[str, str]]) -> str:
    parts = []
    for person_name, image_path in sorted(items, key=lambda x: (x[0], x[1])):
        try:
            stat = os.stat(image_path)
            parts.append(f"{person_name}|{image_path}|{stat.st_size}|{stat.st_mtime}")
        except OSError:
            parts.append(f"{person_name}|{image_path}|missing")
    joined = "\n".join(parts).encode("utf-8", errors="ignore")
    return hashlib.md5(joined).hexdigest()

def person_score_from_embeddings(probe_emb: np.ndarray, ref_embeddings: List[np.ndarray]) -> float:
    if len(ref_embeddings) == 0: return -1.0
    sims = sorted((compute_cosine_similarity(probe_emb, ref_emb) for ref_emb in ref_embeddings), reverse=True)
    top_k = sims[: min(3, len(sims))]
    if len(top_k) == 1: return float(top_k[0])
    return float(0.75 * top_k[0] + 0.25 * (sum(top_k) / len(top_k)))

# ============================================================
# EXCEL ĐIỂM DANH
# ============================================================
def init_excel_file():
    if os.path.exists(EXCEL_FILENAME): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Ho va Ten", "Thoi Gian Diem Danh", "Anh Minh Chung"])
    wb.save(EXCEL_FILENAME)
    print(f"[*] Đã tạo file Excel: {EXCEL_FILENAME}")

def load_logged_names_from_excel():
    global logged_names
    if not os.path.exists(EXCEL_FILENAME): return
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]: logged_names.add(str(row[0]).strip())
        print(f"[*] Đã nạp {len(logged_names)} người đã điểm danh từ Excel.")
    except Exception as e:
        print(f"[!] Lỗi đọc Excel: {e}")

def save_evidence_image(frame: np.ndarray, name: str) -> str:
    ensure_dir(EVIDENCE_DIR)
    timestamp = datetime.now().strftime("%H-%M-%S")
    safe_name = name.replace("/", "_").replace("\\", "_").strip()
    filename = f"{safe_name}_{datetime.now().strftime('%d-%m-%Y')}_{timestamp}.jpg"
    filepath = os.path.join(EVIDENCE_DIR, filename)
    try:
        cv2.imwrite(filepath, frame)
        return filepath
    except Exception as e:
        print(f"[!] Không lưu được ảnh minh chứng cho {name}: {e}")
        return ""

def mark_attendance(name: str, evidence_frame: Optional[np.ndarray] = None) -> bool:
    if name in INVALID_NAMES or name in logged_names: return False
    time_string = datetime.now().strftime("%H:%M:%S")
    evidence_path = ""
    if evidence_frame is not None and getattr(evidence_frame, "size", 0) > 0:
        evidence_path = save_evidence_image(evidence_frame, name)
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        ws.append([name, time_string, evidence_path])
        wb.save(EXCEL_FILENAME)
        logged_names.add(name)
        print(f"\n[+] Điểm danh thành công: {name} lúc {time_string}")
        if evidence_path: print(f"    Ảnh minh chứng: {evidence_path}\n")
        return True
    except Exception as e:
        print(f"[!] Lỗi ghi Excel cho {name}: {e}")
        return False

# ============================================================
# NẠP DỮ LIỆU KHUÔN MẶT
# ============================================================
def get_all_face_images(dataset_dir: str) -> List[Tuple[str, str]]:
    items = []
    for entry in os.listdir(dataset_dir):
        full_path = os.path.join(dataset_dir, entry)
        if os.path.isfile(full_path) and entry.lower().endswith((".jpg", ".jpeg", ".png")):
            name = os.path.splitext(entry)[0].split(" (")[0].strip()
            items.append((name, full_path))
        elif os.path.isdir(full_path):
            person_name = entry.strip()
            for file_name in os.listdir(full_path):
                file_path = os.path.join(full_path, file_name)
                if os.path.isfile(file_path) and file_name.lower().endswith((".jpg", ".jpeg", ".png")):
                    items.append((person_name, file_path))
    return items

def detect_single_face_box(img: np.ndarray, yolo_model) -> Optional[Tuple[int, int, int, int]]:
    try:
        results = yolo_model(img, verbose=False)
        if len(results) == 0 or results[0].boxes is None: return None
        boxes = results[0].boxes
        if len(boxes) != 1: return None
        x1, y1, x2, y2 = map(int, boxes.xyxy[0])
        return (x1, y1, x2, y2)
    except Exception: return None

def crop_best_face(img: np.ndarray, yolo_model) -> Optional[np.ndarray]:
    box = detect_single_face_box(img, yolo_model)
    if box is None: return None
    crop = crop_with_padding(img, box, FACE_PADDING_RATIO)
    if crop.size == 0: return None
    return crop

def build_face_embedding(face_crop: np.ndarray) -> Optional[np.ndarray]:
    try:
        emb_objs = DeepFace.represent(
            img_path=face_crop,
            model_name=EMBED_MODEL_NAME,
            enforce_detection=False,
        )
        if len(emb_objs) == 0: return None
        embedding = np.array(emb_objs[0]["embedding"], dtype=np.float32)
        return l2_normalize(embedding)
    except Exception: return None

def load_known_faces(dataset_path: str, yolo_model, cache_file: str = CACHE_FILE) -> Dict[str, List[np.ndarray]]:
    ensure_dir(dataset_path)
    all_items = get_all_face_images(dataset_path)
    if len(all_items) == 0:
        print(f"[!] Chưa có ảnh mẫu trong thư mục '{dataset_path}'.")
        return {}

    dataset_signature = get_dataset_signature(all_items)

    if os.path.exists(cache_file):
        try:
            print(f"[*] Tìm thấy cache '{cache_file}', đang kiểm tra...")
            with open(cache_file, "rb") as f: data = pickle.load(f)
            if data.get("signature") == dataset_signature and isinstance(data.get("db"), dict):
                print("[*] Cache hợp lệ, dùng lại dữ liệu khuôn mặt.")
                return data["db"]
            print("[*] Cache không còn hợp lệ, sẽ tạo lại.")
        except Exception as e:
            print(f"[!] Lỗi đọc cache: {e}. Sẽ tạo lại cache.")

    print("[*] Đang tạo cơ sở dữ liệu khuôn mặt...")
    person_to_embs: Dict[str, List[np.ndarray]] = {}

    for person_name, image_path in all_items:
        try:
            img = cv2.imread(image_path)
            if img is None: continue

            face_crop = crop_best_face(img, yolo_model)
            if face_crop is None: continue

            h, w = face_crop.shape[:2]
            if h < MIN_FACE_SIZE or w < MIN_FACE_SIZE: continue
            if compute_blur_score(face_crop) < 20.0: continue

            # --- ÁP DỤNG ALIGN & GAMMA CHO DỮ LIỆU GỐC ---
            face_crop = align_face(face_crop)
            face_crop = auto_gamma_correction(face_crop)

            embedding = build_face_embedding(face_crop)
            if embedding is None: continue

            person_to_embs.setdefault(person_name, []).append(embedding)
            print(f"  + OK: {person_name} -> {os.path.basename(image_path)}")

        except Exception as e:
            print(f"  - Bỏ qua '{image_path}': {e}")

    clean_db: Dict[str, List[np.ndarray]] = {}
    for person_name in sorted(person_to_embs.keys()):
        embs = person_to_embs[person_name]
        if len(embs) == 0: continue
        clean_db[person_name] = [l2_normalize(np.asarray(emb, dtype=np.float32)) for emb in embs]

    if clean_db:
        try:
            with open(cache_file, "wb") as f:
                pickle.dump({"db": clean_db, "signature": dataset_signature}, f)
            print(f"[*] Đã lưu cache với {len(clean_db)} người.")
        except Exception as e:
            print(f"[!] Không lưu được cache: {e}")

    return clean_db


# ============================================================
# NHẬN DIỆN 1 KHUÔN MẶT
# ============================================================
def recognize_face(face_crop: np.ndarray, face_w: int) -> Tuple[str, float]:
    if face_crop is None or face_crop.size == 0:
        return "Loi AI", 0.0

    working_crop = face_crop.copy()

    # --- ÁP DỤNG ALIGN & GAMMA CHO LUỒNG CAMERA ---
    try:
        working_crop = align_face(working_crop)
        working_crop = auto_gamma_correction(working_crop)
    except Exception:
        pass

    if face_w < 70:
        try:
            working_crop = cv2.resize(
                working_crop, None, fx=2.0, fy=2.0, interpolation=cv2.INTER_CUBIC
            )
        except Exception:
            pass

    probe_emb = build_face_embedding(working_crop)
    if probe_emb is None or len(known_face_db) == 0:
        return "Unknown", 0.0

    person_scores = {}
    for person_name, ref_embeddings in known_face_db.items():
        person_scores[person_name] = person_score_from_embeddings(probe_emb, ref_embeddings)

    ranked = sorted(person_scores.items(), key=lambda x: x[1], reverse=True)
    if len(ranked) == 0: return "Unknown", 0.0

    best_name, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else None

    threshold = get_face_threshold(face_w)
    gap_threshold = get_gap_threshold(face_w)

    if second_score is None:
        if best_score > threshold + 0.05: return best_name, float(best_score)
        if best_score >= UNKNOWN_MIN_SIM: return "Khong chac", float(best_score)
        return "Loi AI", float(best_score)

    gap = float(best_score - second_score)

    if best_score >= threshold and gap >= gap_threshold:
        return best_name, float(best_score)

    if best_score >= UNKNOWN_MIN_SIM:
        if gap < gap_threshold: return "Khong chac", float(best_score)
        return "Unknown", float(best_score)

    return "Loi AI", float(best_score)


# ============================================================
# TRACKER
# ============================================================
def create_track(box) -> int:
    global next_track_id
    track_id = next_track_id
    next_track_id += 1
    tracks[track_id] = TrackState(box=box)
    return track_id

def update_track(track_id: int, box):
    tracks[track_id].box = box
    tracks[track_id].missing = 0
    tracks[track_id].last_update_time = time.time()

def match_detections_to_tracks(detections: List[Detection], frame_w: int, frame_h: int):
    matched = {}
    unmatched_dets = set(range(len(detections)))
    unmatched_tracks = set(tracks.keys())

    if len(tracks) == 0 or len(detections) == 0:
        return matched, unmatched_dets, unmatched_tracks

    pairs = []
    for tid, tdata in tracks.items():
        tbox = tdata.box
        t_area = box_area(tbox)

        for di, det in enumerate(detections):
            dbox = det.box
            d_area = box_area(dbox)

            iou = compute_iou(tbox, dbox)
            center_dist = normalized_center_distance(tbox, dbox, frame_w, frame_h)
            size_ratio = min(t_area, d_area) / max(t_area, d_area) if max(t_area, d_area) > 0 else 0.0

            if iou < TRACK_MATCH_IOU and center_dist > TRACK_MAX_CENTER_DIST_RATIO:
                continue

            score = (0.60 * iou) + (0.25 * size_ratio) + (0.15 * max(0.0, 1.0 - (center_dist / TRACK_MAX_CENTER_DIST_RATIO)))
            pairs.append((score, tid, di))

    pairs.sort(reverse=True, key=lambda x: x[0])

    used_tracks = set()
    used_dets = set()

    for score, tid, di in pairs:
        if tid in used_tracks or di in used_dets: continue
        matched[tid] = di
        used_tracks.add(tid)
        used_dets.add(di)

    unmatched_dets -= used_dets
    unmatched_tracks -= used_tracks
    return matched, unmatched_dets, unmatched_tracks

def cleanup_tracks(unmatched_tracks):
    remove_ids = []
    for tid in unmatched_tracks:
        if tid not in tracks: continue
        tracks[tid].missing += 1
        if tracks[tid].missing > TRACK_MAX_MISSING:
            remove_ids.append(tid)

    for tid in remove_ids: del tracks[tid]

def update_track_identity(track_id: int, name: str, score: float):
    track = tracks[track_id]

    if name in INVALID_NAMES:
        if track.streak_name is not None:
            if ALLOW_ONE_WEAK_FRAME and track.streak_count >= REQUIRED_CONFIRMATIONS - 1:
                track.score_history.append(score)
                track.score_history = track.score_history[-REQUIRED_CONFIRMATIONS:]
            else:
                track.streak_name = None
                track.streak_count = 0
                track.score_history.clear()
        return

    if track.streak_name == name:
        track.streak_count += 1
    else:
        track.streak_name = name
        track.streak_count = 1
        track.score_history = []

    track.score_history.append(score)
    track.score_history = track.score_history[-REQUIRED_CONFIRMATIONS:]

    avg_score = float(np.mean(track.score_history))
    if avg_score > track.best_score:
        track.best_score = avg_score
        track.best_name = name

    face_w = track.box[2] - track.box[0]
    threshold = get_face_threshold(face_w)

    if track.streak_count >= REQUIRED_CONFIRMATIONS:
        if ALLOW_ONE_WEAK_FRAME:
            weak_frames = sum(s < threshold for s in track.score_history)
            if weak_frames <= 1 and avg_score >= threshold:
                track.confirmed = True
                track.best_name = name
                track.best_score = avg_score
        else:
            all_scores_valid = all(s >= threshold for s in track.score_history)
            if all_scores_valid and avg_score >= threshold:
                track.confirmed = True
                track.best_name = name
                track.best_score = avg_score

def maybe_update_snapshot(track_id: int, frame: np.ndarray, box, recog_score: float):
    track = tracks[track_id]
    evidence = crop_with_padding(frame, box, EVIDENCE_PADDING_RATIO)
    if evidence.size == 0: return

    quality = compute_face_quality(evidence, recog_score)
    if quality > track.snapshot_quality:
        track.snapshot = evidence
        track.snapshot_quality = quality


# ============================================================
# AI WORKER
# ============================================================
def ai_worker():
    global latest_frame, latest_display, ai_processing, ai_fps, is_running

    frame_index = 0

    while is_running:
        if ai_processing:
            time.sleep(0.005)
            continue

        with frame_lock:
            frame_to_process = None if latest_frame is None else latest_frame.copy()

        if frame_to_process is None:
            time.sleep(0.01)
            continue

        ai_processing = True
        start_t = time.time()
        frame_index += 1

        try:
            h, w = frame_to_process.shape[:2]
            entrance_roi = get_entrance_roi(w, h)
            detections: List[Detection] = []

            try:
                if USE_ENTRANCE_ROI:
                    rx1, ry1, rx2, ry2 = entrance_roi
                    detect_region = frame_to_process[ry1:ry2, rx1:rx2].copy()
                    offset_x, offset_y = rx1, ry1
                else:
                    detect_region = frame_to_process.copy()
                    offset_x, offset_y = 0, 0

                detect_region = preprocess_for_detection(detect_region)
                detect_region_big = cv2.resize(
                    detect_region, None, fx=ROI_DETECT_ZOOM, fy=ROI_DETECT_ZOOM, interpolation=cv2.INTER_CUBIC
                )

                results = model(detect_region_big, verbose=False)
                for r in results:
                    if r.boxes is None: continue

                    for i, box in enumerate(r.boxes.xyxy):
                        conf = float(r.boxes.conf[i]) if getattr(r.boxes, "conf", None) is not None else 1.0
                        if conf < YOLO_CONFIDENCE: continue

                        bx1, by1, bx2, by2 = map(int, box)
                        x1 = offset_x + int(bx1 / ROI_DETECT_ZOOM)
                        y1 = offset_y + int(by1 / ROI_DETECT_ZOOM)
                        x2 = offset_x + int(bx2 / ROI_DETECT_ZOOM)
                        y2 = offset_y + int(by2 / ROI_DETECT_ZOOM)

                        x1, y1 = max(0, x1), max(0, y1)
                        x2, y2 = min(w, x2), min(h, y2)

                        face_w, face_h = x2 - x1, y2 - y1
                        if face_w <= 0 or face_h <= 0: continue

                        cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
                        if USE_ENTRANCE_ROI and not point_in_box(cx, cy, entrance_roi): continue

                        if face_w < MIN_FACE_SIZE or face_h < MIN_FACE_SIZE:
                            detections.append(Detection(box=(x1, y1, x2, y2), confidence=conf, name="Qua xa", score=0.0, need_recognition=False))
                        else:
                            detections.append(Detection(box=(x1, y1, x2, y2), confidence=conf))

            except Exception as e:
                print(f"[!] Lỗi detect: {e}")

            matched, unmatched_dets, unmatched_tracks = match_detections_to_tracks(detections, w, h)
            track_to_det: Dict[int, Detection] = {}

            for tid, det_idx in matched.items():
                det = detections[det_idx]
                update_track(tid, det.box)
                track_to_det[tid] = det

            for det_idx in unmatched_dets:
                det = detections[det_idx]
                tid = create_track(det.box)
                track_to_det[tid] = det

            cleanup_tracks(unmatched_tracks)

            for tid, det in track_to_det.items():
                x1, y1, x2, y2 = det.box
                face_w = x2 - x1
                track = tracks[tid]

                if not det.need_recognition:
                    track.best_name = det.name
                    track.best_score = det.score
                    continue

                if track.attendance_marked: continue
                if frame_index - track.last_recognized_frame < RECOGNIZE_EVERY_N_FRAMES: continue

                face_crop = crop_with_padding(frame_to_process, det.box, FACE_PADDING_RATIO)
                if face_crop.size == 0: continue

                track.last_recognized_frame = frame_index
                name, score = recognize_face(face_crop, face_w)
                update_track_identity(tid, name, score)

                if name not in INVALID_NAMES:
                    maybe_update_snapshot(tid, frame_to_process, det.box, score)

                if track.confirmed and not track.attendance_marked:
                    confirmed_name = track.best_name
                    if confirmed_name not in logged_names:
                        evidence = track.snapshot if track.snapshot is not None else face_crop
                        success = mark_attendance(confirmed_name, evidence)
                        if success: track.attendance_marked = True

            display_items = []
            for tid, tdata in tracks.items():
                x1, y1, x2, y2 = tdata.box
                show_name = tdata.best_name
                show_score = tdata.best_score

                if tdata.confirmed: color = (0, 200, 0)
                elif show_name in INVALID_NAMES: color = (0, 0, 255)
                else: color = (0, 165, 255)

                display_items.append({
                    "track_id": tid, "box": (x1, y1, x2, y2),
                    "name": show_name, "score": show_score,
                    "color": color, "confirmed": tdata.confirmed,
                })

            with result_lock:
                latest_display = display_items

            elapsed = time.time() - start_t
            ai_fps = 1.0 / elapsed if elapsed > 0 else 0.0

        except Exception as e:
            print(f"[!] Lỗi AI worker: {e}")
            traceback.print_exc()

        finally:
            ai_processing = False


# ============================================================
# MAIN
# ============================================================
def main():
    global model, latest_frame, is_running, known_face_db

    print("\n" + "=" * 60)
    print("HỆ THỐNG ĐIỂM DANH TỰ ĐỘNG - BẢN TỐI ƯU UX (ALIGN & GAMMA)")
    print("=" * 60)

    ensure_dir(KNOWN_FACES_DIR)
    ensure_dir(EVIDENCE_DIR)
    init_excel_file()
    load_logged_names_from_excel()

    print("[*] Đang nạp YOLO face detector...")
    try:
        model = YOLO(YOLO_MODEL_PATH)
    except Exception as e:
        print(f"[!] Không thể nạp model YOLO: {e}")
        return

    known_face_db = load_known_faces(KNOWN_FACES_DIR, model)
    if len(known_face_db) == 0:
        print("[!] Chưa có dữ liệu khuôn mặt mẫu hợp lệ trong 'known_faces'.")
        return

    total_templates = sum(len(v) for v in known_face_db.values())
    print(f"[*] Đã nạp {len(known_face_db)} người với {total_templates} ảnh mẫu hợp lệ.")

    cap = cv2.VideoCapture(CAMERA_INDEX)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_WIDTH)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)

    if not cap.isOpened():
        print("[!] Không mở được webcam.")
        return

    actual_w = cap.get(cv2.CAP_PROP_FRAME_WIDTH)
    actual_h = cap.get(cv2.CAP_PROP_FRAME_HEIGHT)
    print(f"[*] Camera actual resolution: {actual_w} x {actual_h}")

    ai_thread = threading.Thread(target=ai_worker, daemon=True)
    ai_thread.start()

    prev_t = time.time()

    print("\n[*] Hệ thống sẵn sàng. Nhấn ESC để thoát.")
    print("[*] Nếu ROI chưa đúng vị trí cửa lớp, hãy chỉnh get_entrance_roi().\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[!] Mất kết nối camera.")
            break

        fh, fw = frame.shape[:2]
        entrance_roi = get_entrance_roi(fw, fh)

        with frame_lock:
            latest_frame = frame.copy()

        curr_t = time.time()
        cam_fps = 1.0 / max(1e-6, (curr_t - prev_t))
        prev_t = curr_t

        if USE_ENTRANCE_ROI:
            rx1, ry1, rx2, ry2 = entrance_roi
            cv2.rectangle(frame, (rx1, ry1), (rx2, ry2), (255, 200, 0), 2)
            cv2.putText(
                frame, "Vung diem danh", (rx1, max(20, ry1 - 10)),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 200, 0), 2,
            )

        with result_lock:
            faces_to_draw = latest_display.copy()

        for item in faces_to_draw:
            x1, y1, x2, y2 = item["box"]
            name = item["name"]
            score = item["score"]
            color = item["color"]
            confirmed = item["confirmed"]

            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            label = make_label(name, score)
            if confirmed: label += " | OK"

            (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.55, 2)
            y_top = max(0, y1 - th - 8)
            cv2.rectangle(frame, (x1, y_top), (x1 + tw + 6, y1), color, -1)
            cv2.putText(
                frame, label, (x1 + 3, y1 - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 2,
            )

        cv2.putText(frame, f"Cam FPS: {cam_fps:.1f}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 0), 2)
        cv2.putText(frame, f"AI FPS: {ai_fps:.1f}", (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 200, 255), 2)
        cv2.putText(frame, f"So nguoi da diem danh: {len(logged_names)}", (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 255, 0), 2)

        if SHOW_DEBUG:
            cv2.putText(frame, f"Tracks: {len(tracks)}", (10, 120), cv2.FONT_HERSHEY_SIMPLEX, 0.60, (180, 255, 180), 2)

        cv2.imshow(WINDOW_NAME, frame)
        key = cv2.waitKey(1) & 0xFF
        if key == 27:
            is_running = False
            break

    cap.release()
    cv2.destroyAllWindows()
    ai_thread.join(timeout=1.0)
    print("\n[!] Đã tắt hệ thống.")


if __name__ == "__main__":
    main()