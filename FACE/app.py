import cv2
import os
import time
import pickle
import threading
import numpy as np
from datetime import datetime
from ultralytics import YOLO
from deepface import DeepFace
from cv2 import dnn_superres
from openpyxl import Workbook, load_workbook
from flask import Flask, Response, jsonify, render_template_string

# ============================================================
# CẤU HÌNH HỆ THỐNG
# ============================================================
CAMERA_INDEX = 0
YOLO_MODEL_PATH = "yolov8n-face.pt"
FSRCNN_MODEL_PATH = "FSRCNN_x4.pb"
KNOWN_FACES_DIR = "known_faces"
CACHE_FILE = "face_cache.pkl"

USE_FSRCNN = True
USE_ENTRANCE_ROI = True

YOLO_CONFIDENCE = 0.55
MIN_FACE_SIZE = 45

MATCH_THRESHOLD_NEAR = 0.55
MATCH_THRESHOLD_FAR = 0.48
GAP_THRESHOLD = 0.03

REQUIRED_CONFIRMATIONS = 3
RECOGNIZE_EVERY_N_FRAMES = 3
TRACK_MAX_MISSING = 12
TRACK_MATCH_IOU = 0.30

EXCEL_FILENAME = f"DanhSach_DiemDanh_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
EVIDENCE_DIR = "attendance_evidence"

# ============================================================
# BIẾN TOÀN CỤC & FLASK APP
# ============================================================
app = Flask(__name__)

latest_frame = None
annotated_frame = None 
display_faces = []
is_running = True

frame_lock = threading.Lock()
result_lock = threading.Lock()

ai_processing = False
ai_fps = 0.0
cam_fps = 0.0

logged_names = set()
attendance_web_list = [] 

tracks = {}
next_track_id = 0

known_face_encodings = []
known_face_names = []

model = None
sr = None

# ============================================================
# GIAO DIỆN WEB (Đã tích hợp API thực)
# ============================================================
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Hệ Thống Điểm Danh AI</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        ::-webkit-scrollbar { width: 6px; }
        ::-webkit-scrollbar-track { background: #f1f5f9; border-radius: 8px; }
        ::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 8px; }
        ::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
        @keyframes slideIn {
            from { opacity: 0; transform: translateY(-10px); }
            to { opacity: 1; transform: translateY(0); }
        }
    </style>
</head>
<body class="bg-slate-100 h-screen overflow-hidden flex flex-col font-sans text-slate-800">
    <!-- Header -->
    <header class="bg-gradient-to-r from-blue-800 to-blue-600 text-white shadow-lg p-3 lg:p-4 flex justify-between items-center z-10 shrink-0">
        <div class="flex items-center gap-3 lg:gap-4">
            <div class="h-10 w-10 bg-white/10 backdrop-blur rounded-full flex items-center justify-center text-white border border-white/20 shadow-inner">
                <i class="fa-solid fa-expand"></i>
            </div>
            <div>
                <h1 class="text-lg lg:text-xl font-bold tracking-wider uppercase text-white drop-shadow-sm">Hệ Thống Điểm Danh Thông Minh</h1>
                <p class="text-xs text-blue-200 font-medium">Đồ án: Lê Gia Huy</p>
            </div>
        </div>
        <div class="hidden sm:flex gap-3 lg:gap-4 text-sm font-semibold">
            <div class="bg-blue-900/50 backdrop-blur px-3 lg:px-4 py-1.5 rounded-lg border border-blue-400/30 shadow-inner flex items-center gap-2">
                <i class="fa-solid fa-camera text-slate-300"></i> Cam FPS: <span id="cam-fps" class="text-yellow-400 font-mono text-base lg:text-lg">0.0</span>
            </div>
            <div class="bg-blue-900/50 backdrop-blur px-3 lg:px-4 py-1.5 rounded-lg border border-blue-400/30 shadow-inner flex items-center gap-2">
                <i class="fa-solid fa-microchip text-slate-300"></i> AI FPS: <span id="ai-fps" class="text-green-400 font-mono text-base lg:text-lg">0.0</span>
            </div>
        </div>
    </header>

    <!-- Main Content -->
    <main class="flex-1 flex flex-col lg:flex-row gap-4 p-4 overflow-hidden h-full">
        <!-- Left: Video Stream -->
        <div class="flex-[2] bg-white rounded-xl shadow-md border border-slate-200 overflow-hidden flex flex-col relative h-full">
            <div class="bg-slate-800 text-slate-200 py-2.5 px-4 text-sm font-semibold flex justify-between items-center shadow-md z-10">
                <span class="flex items-center"><i class="fa-solid fa-circle text-red-500 mr-2 animate-pulse text-xs"></i>Camera Trực Tiếp</span>
                <span class="text-[10px] sm:text-xs bg-slate-700/50 px-2 py-1 rounded text-green-400 border border-slate-600 flex items-center">
                    <i class="fa-solid fa-shield-halved mr-1"></i> IoU Tracker Active
                </span>
            </div>
            
            <div class="flex-1 bg-black relative flex items-center justify-center overflow-hidden">
                <!-- Video Stream từ Flask Backend -->
                <img src="/video_feed" class="w-full h-full object-contain" alt="Video Stream Đang Khởi Động...">
                
                <!-- Status overlay -->
                <div class="absolute bottom-4 left-4 bg-black/60 backdrop-blur-sm text-white px-3 py-1.5 rounded shadow text-xs font-mono flex items-center border border-white/10">
                    <i class="fa-solid fa-circle text-red-500 text-[8px] mr-2 animate-pulse"></i> REC • YOLOv8 + ArcFace
                </div>
            </div>
        </div>

        <!-- Right: Log Panel -->
        <div class="flex-[1] lg:max-w-[350px] bg-white rounded-xl shadow-md border border-slate-200 flex flex-col overflow-hidden h-full">
            <div class="bg-gradient-to-r from-slate-50 to-white border-b border-slate-200 py-4 px-5 shrink-0">
                <h2 class="text-base font-bold text-slate-800 flex items-center">
                    <div class="bg-blue-100 text-blue-600 p-1.5 rounded mr-2"><i class="fa-solid fa-list-check"></i></div>
                    Lịch Sử Điểm Danh
                </h2>
                <div class="flex justify-between items-end mt-2">
                    <p class="text-[11px] font-medium text-slate-500 uppercase tracking-wider">
                        <i class="fa-regular fa-calendar-days mr-1"></i> <span id="current-date"></span>
                    </p>
                    <span class="bg-emerald-100 text-emerald-700 border border-emerald-200 px-2 py-0.5 rounded text-[9px] font-bold uppercase flex items-center gap-1 shadow-sm">
                        <span class="w-1.5 h-1.5 bg-emerald-500 rounded-full animate-pulse"></span> Live
                    </span>
                </div>
            </div>
            
            <div class="flex-1 overflow-y-auto p-3 bg-slate-50/50" id="attendance-list">
                <!-- Data from API -->
            </div>
            
            <div class="bg-white border-t border-slate-200 p-4 shrink-0 flex justify-between items-center shadow-[0_-4px_10px_-5px_rgba(0,0,0,0.05)]">
                <div class="flex flex-col">
                    <span class="text-[10px] text-slate-400 uppercase font-bold tracking-widest mb-0.5">Tổng số người</span>
                    <div class="flex items-baseline gap-1">
                        <span class="text-2xl font-black text-blue-700 leading-none" id="total-count">0</span>
                    </div>
                </div>
            </div>
        </div>
    </main>

    <script>
        const dateOptions = { day: '2-digit', month: '2-digit', year: 'numeric' };
        document.getElementById('current-date').innerText = new Date().toLocaleDateString('vi-VN', dateOptions);

        let previousCount = 0;

        async function fetchData() {
            try {
                const response = await fetch('/api/data');
                const data = await response.json();
                
                document.getElementById('cam-fps').innerText = data.cam_fps.toFixed(1);
                document.getElementById('ai-fps').innerText = data.ai_fps.toFixed(1);
                document.getElementById('total-count').innerText = data.attendance.length;

                const listContainer = document.getElementById('attendance-list');
                
                if(data.attendance.length === 0) {
                    listContainer.innerHTML = '<div class="h-full flex flex-col items-center justify-center text-slate-400"><div class="w-16 h-16 bg-slate-100 rounded-full flex items-center justify-center mb-3 border border-slate-200"><i class="fa-solid fa-clipboard-user text-2xl text-slate-300"></i></div><p class="font-medium text-sm text-slate-500">Chưa có dữ liệu</p><p class="text-xs text-slate-400 mt-1">Đang chờ quét khuôn mặt...</p></div>';
                    previousCount = 0;
                    return;
                }

                // Nếu có dữ liệu mới, render lại danh sách
                if(data.attendance.length !== previousCount) {
                    let html = '';
                    data.attendance.slice().reverse().forEach((item, index) => {
                        const animateClass = (index === 0 && data.attendance.length > previousCount) ? 'animate-[slideIn_0.3s_ease-out]' : '';
                        html += `
                        <div class="flex items-center gap-3 p-2.5 mb-2 rounded-lg border border-slate-200 bg-white hover:border-blue-300 shadow-sm ${animateClass}">
                            <div class="h-9 w-9 shrink-0 rounded-full bg-emerald-100 text-emerald-600 border border-emerald-200 flex items-center justify-center font-bold">
                                <i class="fa-solid fa-check text-sm"></i>
                            </div>
                            <div class="flex-1 min-w-0">
                                <h3 class="font-bold text-slate-700 text-sm truncate">${item.name}</h3>
                                <div class="flex items-center gap-2 mt-0.5">
                                    <span class="text-[10px] font-mono text-slate-500 flex items-center"><i class="fa-regular fa-clock mr-1"></i>${item.time}</span>
                                </div>
                            </div>
                            <div class="shrink-0 text-[10px] text-emerald-600 font-bold bg-emerald-50 px-1.5 py-0.5 rounded border border-emerald-100 uppercase">Thành công</div>
                        </div>`;
                    });
                    listContainer.innerHTML = html;
                    previousCount = data.attendance.length;
                }
            } catch (error) { console.error("Lỗi cập nhật dữ liệu:", error); }
        }

        setInterval(fetchData, 1000);
        fetchData();
    </script>
</body>
</html>
"""

# ============================================================
# CÁC HÀM XỬ LÝ NHƯ BẢN GỐC
# ============================================================
def ensure_dir(path):
    if not os.path.exists(path):
        os.makedirs(path)

def compute_cosine_similarity(feat1, feat2):
    n1 = np.linalg.norm(feat1)
    n2 = np.linalg.norm(feat2)
    if n1 < 1e-8 or n2 < 1e-8: return -1.0
    return float(np.dot(feat1, feat2) / (n1 * n2))

def compute_iou(boxA, boxB):
    ax1, ay1, ax2, ay2 = boxA
    bx1, by1, bx2, by2 = boxB
    inter_w = max(0, min(ax2, bx2) - max(ax1, bx1))
    inter_h = max(0, min(ay2, by2) - max(ay1, by1))
    inter_area = inter_w * inter_h
    union = max(0, ax2 - ax1) * max(0, ay2 - ay1) + max(0, bx2 - bx1) * max(0, by2 - by1) - inter_area
    return inter_area / union if union > 0 else 0.0

def point_in_box(cx, cy, box):
    x1, y1, x2, y2 = box
    return x1 <= cx <= x2 and y1 <= cy <= y2

def get_face_threshold(face_w):
    return MATCH_THRESHOLD_FAR if face_w < 75 else MATCH_THRESHOLD_NEAR

def make_label(name, score):
    return f"{name} ({score*100:.0f}%)" if score > 0 else name

def get_entrance_roi(frame_w, frame_h):
    roi_w = int(frame_w * 0.75)
    roi_h = int(frame_h * 0.85)
    x1 = (frame_w - roi_w) // 2
    y1 = (frame_h - roi_h) // 2
    return (x1, y1, x1 + roi_w, y1 + roi_h)

def init_excel_file():
    if os.path.exists(EXCEL_FILENAME): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Ho va Ten", "Thoi Gian Diem Danh", "Anh Minh Chung"])
    wb.save(EXCEL_FILENAME)

def load_logged_names_from_excel():
    global logged_names, attendance_web_list
    if not os.path.exists(EXCEL_FILENAME): return
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[0] != "Ho va Ten":
                name = str(row[0]).strip()
                time_str = str(row[1]).strip() if row[1] else ""
                logged_names.add(name)
                attendance_web_list.append({"name": name, "time": time_str})
    except Exception as e:
        print(f"[!] Lỗi đọc Excel: {e}")

def save_evidence_image(frame, name):
    ensure_dir(EVIDENCE_DIR)
    timestamp = datetime.now().strftime("%H-%M-%S")
    safe_name = name.replace("/", "_").replace("\\", "_").strip()
    filepath = os.path.join(EVIDENCE_DIR, f"{safe_name}_{datetime.now().strftime('%d-%m-%Y')}_{timestamp}.jpg")
    try:
        cv2.imwrite(filepath, frame)
        return filepath
    except: return ""

def mark_attendance(name, evidence_frame=None):
    if name in {"Unknown", "Qua xa", "Loi AI", "Khong chac"} or name in logged_names:
        return False
    time_string = datetime.now().strftime("%H:%M:%S")
    evidence_path = save_evidence_image(evidence_frame, name) if evidence_frame is not None else ""
    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        ws.append([name, time_string, evidence_path])
        wb.save(EXCEL_FILENAME)
        
        logged_names.add(name)
        attendance_web_list.append({"name": name, "time": time_string})
        print(f"\n[+] TING TING! Điểm danh thành công: {name} lúc {time_string}")
        return True
    except Exception as e:
        print(f"[!] Lỗi ghi Excel cho {name}: {e}")
        return False

def get_all_face_images(dataset_dir):
    items = []
    for entry in os.listdir(dataset_dir):
        full_path = os.path.join(dataset_dir, entry)
        if os.path.isfile(full_path) and entry.lower().endswith((".jpg", ".jpeg", ".png")):
            items.append((os.path.splitext(entry)[0].split(" (")[0].strip(), full_path))
        elif os.path.isdir(full_path):
            person_name = entry.strip()
            for file_name in os.listdir(full_path):
                if file_name.lower().endswith((".jpg", ".jpeg", ".png")):
                    items.append((person_name, os.path.join(full_path, file_name)))
    return items

def crop_best_face(img, yolo_model):
    try:
        results = yolo_model(img, verbose=False)
        boxes = results[0].boxes
        if boxes is None or len(boxes) == 0: return None
        best_box = max(boxes.xyxy, key=lambda b: (b[2] - b[0]) * (b[3] - b[1]))
        x1, y1, x2, y2 = map(int, best_box)
        h, w = img.shape[:2]
        pad_w, pad_h = int((x2 - x1) * 0.20), int((y2 - y1) * 0.20)
        face_crop = img[max(0, y1-pad_h):min(h, y2+pad_h), max(0, x1-pad_w):min(w, x2+pad_w)]
        return face_crop if face_crop.size > 0 else None
    except: return None

def load_known_faces(dataset_path, yolo_model, cache_file=CACHE_FILE):
    known_encodings, known_names = [], []
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "rb") as f: data = pickle.load(f)
            return data["encodings"], data["names"]
        except: pass
    ensure_dir(dataset_path)
    all_items = get_all_face_images(dataset_path)
    for person_name, image_path in all_items:
        try:
            img = cv2.imread(image_path)
            face_crop = crop_best_face(img, yolo_model)
            if face_crop is None: continue
            emb_objs = DeepFace.represent(img_path=face_crop, model_name="ArcFace", enforce_detection=False)
            if len(emb_objs) > 0:
                known_encodings.append(np.array(emb_objs[0]["embedding"], dtype=np.float32))
                known_names.append(person_name)
        except: pass
    if known_encodings:
        try:
            with open(cache_file, "wb") as f: pickle.dump({"encodings": known_encodings, "names": known_names}, f)
        except: pass
    return known_encodings, known_names

def recognize_face(face_crop, face_w):
    if face_crop is None or face_crop.size == 0: return "Loi AI", 0.0
    working_crop = face_crop.copy()
    if USE_FSRCNN and face_w < 60 and sr is not None:
        try: working_crop = sr.upsample(working_crop)
        except: pass
    try:
        emb_objs = DeepFace.represent(img_path=working_crop, model_name="ArcFace", enforce_detection=False)
        if len(emb_objs) == 0 or len(known_face_encodings) == 0: return "Unknown", 0.0
        embedding = np.array(emb_objs[0]["embedding"], dtype=np.float32)
        similarities = [compute_cosine_similarity(embedding, known_emb) for known_emb in known_face_encodings]
        sorted_indices = np.argsort(similarities)[::-1]
        best_idx = sorted_indices[0]
        best_sim = similarities[best_idx]
        second_sim = -1.0
        for idx in sorted_indices[1:]:
            if known_face_names[idx] != known_face_names[best_idx]:
                second_sim = similarities[idx]
                break
        threshold = get_face_threshold(face_w)
        if best_sim > threshold:
            if (best_sim - second_sim) > GAP_THRESHOLD: return known_face_names[best_idx], best_sim
            return "Khong chac", best_sim
        return "Unknown", best_sim
    except: return "Loi AI", 0.0

# Tracker functions
def create_track(box):
    global next_track_id
    tid = next_track_id
    next_track_id += 1
    tracks[tid] = {"box": box, "missing": 0, "name_votes": {}, "best_name": "Unknown", "best_score": 0.0, "confirmed": False, "attendance_marked": False, "snapshot": None}
    return tid

def update_track(track_id, box):
    tracks[track_id]["box"] = box
    tracks[track_id]["missing"] = 0

def match_detections_to_tracks(detections):
    matched = {}
    unmatched_dets, unmatched_tracks = set(range(len(detections))), set(tracks.keys())
    if not tracks: return matched, unmatched_dets, unmatched_tracks
    pairs = []
    for tid, tdata in tracks.items():
        for di, det in enumerate(detections):
            iou = compute_iou(tdata["box"], det["box"])
            if iou >= TRACK_MATCH_IOU: pairs.append((iou, tid, di))
    pairs.sort(reverse=True, key=lambda x: x[0])
    used_tracks, used_dets = set(), set()
    for iou, tid, di in pairs:
        if tid in used_tracks or di in used_dets: continue
        matched[tid] = di
        used_tracks.add(tid)
        used_dets.add(di)
    return matched, unmatched_dets - used_dets, unmatched_tracks - used_tracks

def cleanup_tracks():
    remove_ids = []
    for tid, data in tracks.items():
        data["missing"] = data.get("missing", 0) + 1
        if data["missing"] > TRACK_MAX_MISSING:
            remove_ids.append(tid)
    for tid in remove_ids:
        del tracks[tid]

def update_track_identity(track_id, name, score):
    if name in {"Unknown", "Qua xa", "Loi AI", "Khong chac"}: return
    track = tracks[track_id]
    track["name_votes"][name] = track["name_votes"].get(name, 0) + 1
    if score > track["best_score"]:
        track["best_score"], track["best_name"] = score, name
    if track["name_votes"][name] >= REQUIRED_CONFIRMATIONS:
        track["confirmed"], track["best_name"] = True, name

# ============================================================
# LUỒNG CAMERA VÀ AI DÀNH CHO WEB
# ============================================================
def camera_worker():
    global latest_frame, annotated_frame, cam_fps, is_running
    cap = cv2.VideoCapture(CAMERA_INDEX)
    prev_t = time.time()
    
    while is_running:
        ret, frame = cap.read()
        if not ret: 
            time.sleep(0.1)
            continue

        fh, fw = frame.shape[:2]
        entrance_roi = get_entrance_roi(fw, fh)

        with frame_lock:
            latest_frame = frame.copy()

        curr_t = time.time()
        cam_fps = 1.0 / (curr_t - prev_t) if (curr_t - prev_t) > 0 else 0.0
        prev_t = curr_t

        # Vẽ đồ họa lên khung hình để phát trực tiếp lên Web
        draw_frame = frame.copy()
        if USE_ENTRANCE_ROI:
            rx1, ry1, rx2, ry2 = entrance_roi
            cv2.rectangle(draw_frame, (rx1, ry1), (rx2, ry2), (255, 200, 0), 2)
            cv2.putText(draw_frame, "Vung Kiem Soat An Ninh", (rx1, max(20, ry1 - 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 200, 0), 2)

        with result_lock:
            faces_to_draw = display_faces.copy()

        for item in faces_to_draw:
            x1, y1, x2, y2 = item["box"]
            color = item["color"]
            label = f"{make_label(item['name'], item['score'])}"
            if item["confirmed"]: label += " | DA XAC NHAN"

            cv2.rectangle(draw_frame, (x1, y1), (x2, y2), color, 2)
            (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.55, 2)
            y_top = max(0, y1 - th - 8)
            cv2.rectangle(draw_frame, (x1, y_top), (x1 + tw + 6, y1), color, -1)
            cv2.putText(draw_frame, label, (x1 + 3, y1 - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 2)

        annotated_frame = draw_frame
        
    cap.release()

def ai_worker():
    global display_faces, ai_processing, ai_fps, is_running
    frame_index = 0

    while is_running:
        if ai_processing:
            time.sleep(0.005)
            continue

        with frame_lock:
            frame_to_process = None if latest_frame is None else latest_frame.copy()

        if frame_to_process is None:
            time.sleep(0.01)
            continue

        ai_processing = True
        start_t = time.time()
        frame_index += 1

        h, w = frame_to_process.shape[:2]
        entrance_roi = get_entrance_roi(w, h)
        detections = []

        try:
            results = model(frame_to_process, verbose=False)
            for r in results:
                if r.boxes is None: continue
                for i, box in enumerate(r.boxes.xyxy):
                    conf = float(r.boxes.conf[i]) if getattr(r.boxes, "conf", None) is not None else 1.0
                    if conf < YOLO_CONFIDENCE: continue
                    x1, y1, x2, y2 = map(int, box)
                    x1, y1, x2, y2 = max(0, x1), max(0, y1), min(w, x2), min(h, y2)
                    face_w, face_h = x2 - x1, y2 - y1
                    if face_w <= 0 or face_h <= 0: continue
                    cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
                    if USE_ENTRANCE_ROI and not point_in_box(cx, cy, entrance_roi): continue
                    if face_w < MIN_FACE_SIZE or face_h < MIN_FACE_SIZE:
                        detections.append({"box": (x1, y1, x2, y2), "name": "Qua xa", "score": 0.0, "need_recognition": False})
                        continue
                    detections.append({"box": (x1, y1, x2, y2), "name": "Unknown", "score": 0.0, "need_recognition": True})
        except: pass

        matched, unmatched_dets, unmatched_tracks = match_detections_to_tracks(detections)
        for tid, det_idx in matched.items(): update_track(tid, detections[det_idx]["box"])
        for det_idx in unmatched_dets: create_track(detections[det_idx]["box"])
        cleanup_tracks()

        track_to_det = {}
        for tid, det_idx in matched.items(): track_to_det[tid] = detections[det_idx]
        for tid, tdata in tracks.items():
            if tid not in track_to_det:
                for det in detections:
                    if det["box"] == tdata["box"]:
                        track_to_det[tid] = det
                        break

        for tid, det in track_to_det.items():
            x1, y1, x2, y2 = det["box"]
            face_w, face_h = x2 - x1, y2 - y1
            if tracks[tid]["snapshot"] is None:
                pad = 20
                tracks[tid]["snapshot"] = frame_to_process[max(0, y1-pad):min(h, y2+pad), max(0, x1-pad):min(w, x2+pad)].copy()
            if not det["need_recognition"]:
                tracks[tid]["best_name"], tracks[tid]["best_score"] = det["name"], det["score"]
                continue
            if frame_index % RECOGNIZE_EVERY_N_FRAMES != 0: continue

            pad_w, pad_h = int(face_w * 0.20), int(face_h * 0.20)
            face_crop = frame_to_process[max(0, y1-pad_h):min(h, y2+pad_h), max(0, x1-pad_w):min(w, x2+pad_w)]
            if face_crop.size == 0: continue

            name, score = recognize_face(face_crop, face_w)
            update_track_identity(tid, name, score)

            if tracks[tid]["confirmed"] and not tracks[tid]["attendance_marked"]:
                confirmed_name = tracks[tid]["best_name"]
                if confirmed_name not in logged_names:
                    evidence = tracks[tid]["snapshot"] if tracks[tid]["snapshot"] is not None else frame_to_process
                    if mark_attendance(confirmed_name, evidence):
                        tracks[tid]["attendance_marked"] = True

        temp_display = []
        for tid, tdata in tracks.items():
            show_name = tdata["best_name"]
            if tdata["confirmed"]: color = (0, 200, 0)
            elif show_name in ["Unknown", "Khong chac", "Qua xa", "Loi AI"]: color = (0, 0, 255)
            else: color = (0, 165, 255)
            temp_display.append({"track_id": tid, "box": tdata["box"], "name": show_name, "score": tdata["best_score"], "color": color, "confirmed": tdata["confirmed"]})

        with result_lock: display_faces = temp_display
        elapsed = time.time() - start_t
        ai_fps = 1.0 / elapsed if elapsed > 0 else 0.0
        ai_processing = False

# ============================================================
# CÁC ĐỊA CHỈ API CỦA WEB FLASK
# ============================================================
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

def generate_video_stream():
    global annotated_frame
    while True:
        if annotated_frame is None:
            time.sleep(0.05)
            continue
        # Chuyển đổi khung hình OpenCV thành JPEG để web hiển thị mượt mà
        ret, buffer = cv2.imencode('.jpg', annotated_frame)
        if not ret: continue
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

@app.route('/video_feed')
def video_feed():
    return Response(generate_video_stream(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/data')
def get_data():
    # Gửi dữ liệu tĩnh (FPS, Danh sách) lên giao diện bằng định dạng JSON
    return jsonify({
        "cam_fps": cam_fps,
        "ai_fps": ai_fps,
        "attendance": attendance_web_list
    })

# ============================================================
# KHỞI CHẠY HỆ THỐNG
# ============================================================
if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("HE THONG DIEM DANH AI - TRUNG TAM MAY CHU WEB")
    print("Do an: Le Gia Huy")
    print("=" * 60)

    ensure_dir(KNOWN_FACES_DIR)
    ensure_dir(EVIDENCE_DIR)
    init_excel_file()
    load_logged_names_from_excel()

    print("[*] Đang nạp hệ thống nhận diện YOLOv8...")
    try:
        model = YOLO(YOLO_MODEL_PATH)
    except Exception as e:
        print(f"[!] Không thể nạp model YOLO: {e}")
        exit()

    print("[*] Đang khởi tạo kính lúp AI FSRCNN...")
    try:
        sr_obj = dnn_superres.DnnSuperResImpl_create()
        sr_obj.readModel(FSRCNN_MODEL_PATH)
        sr_obj.setModel("fsrcnn", 4)
        sr = sr_obj
        print("[*] FSRCNN sẵn sàng.")
    except:
        sr = None
        print("[!] Không tìm thấy FSRCNN, tắt upscale.")

    known_face_encodings, known_face_names = load_known_faces(KNOWN_FACES_DIR, model)
    print(f"[*] Đã nạp {len(known_face_encodings)} khuôn mặt.")

    # Khởi chạy luồng Camera và AI chạy ngầm
    threading.Thread(target=camera_worker, daemon=True).start()
    threading.Thread(target=ai_worker, daemon=True).start()

    print("\n" + "*"*50)
    print("[+] MÁY CHỦ ĐÃ SẴN SÀNG!")
    print("[+] BẠN HÃY MỞ TRÌNH DUYỆT (CHROME/CỐC CỐC) VÀ TRUY CẬP VÀO ĐƯỜNG LINK NÀY:")
    print("    👉  http://127.0.0.1:5000  👈")
    print("*"*50 + "\n")

    # Chạy Web Server (Tắt reloader để không bị xung đột Camera)
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)