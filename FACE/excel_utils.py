import os
import cv2
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from typing import Optional

from config import EXCEL_FILENAME, EVIDENCE_DIR, INVALID_NAMES
from utils import ensure_dir

def init_excel_file():
    if os.path.exists(EXCEL_FILENAME):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Ho va Ten", "Thoi Gian Diem Danh", "Anh Minh Chung"])
    wb.save(EXCEL_FILENAME)
    print(f"[*] Đã tạo file Excel: {EXCEL_FILENAME}")

def load_logged_names_from_excel(logged_names_set: set):
    if not os.path.exists(EXCEL_FILENAME):
        return

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                logged_names_set.add(str(row[0]).strip())
        print(f"[*] Đã nạp {len(logged_names_set)} người đã điểm danh từ Excel.")
    except Exception as e:
        print(f"[!] Lỗi đọc Excel: {e}")

def save_evidence_image(frame: np.ndarray, name: str) -> str:
    ensure_dir(EVIDENCE_DIR)
    timestamp = datetime.now().strftime("%H-%M-%S")
    safe_name = name.replace("/", "_").replace("\\", "_").strip()
    filename = f"{safe_name}_{datetime.now().strftime('%d-%m-%Y')}_{timestamp}.jpg"
    filepath = os.path.join(EVIDENCE_DIR, filename)

    try:
        cv2.imwrite(filepath, frame)
        return filepath
    except Exception as e:
        print(f"[!] Không lưu được ảnh minh chứng cho {name}: {e}")
        return ""

def mark_attendance(name: str, logged_names_set: set, evidence_frame: Optional[np.ndarray] = None) -> bool:
    if name in INVALID_NAMES or name in logged_names_set:
        return False

    time_string = datetime.now().strftime("%H:%M:%S")
    evidence_path = ""

    if evidence_frame is not None and getattr(evidence_frame, "size", 0) > 0:
        evidence_path = save_evidence_image(evidence_frame, name)

    try:
        wb = load_workbook(EXCEL_FILENAME)
        ws = wb.active
        ws.append([name, time_string, evidence_path])
        wb.save(EXCEL_FILENAME)

        logged_names_set.add(name)
        print(f"\n[+] Điểm danh thành công: {name} lúc {time_string}")
        if evidence_path:
            print(f"    Ảnh minh chứng: {evidence_path}\n")
        return True
    except Exception as e:
        print(f"[!] Lỗi ghi Excel cho {name}: {e}")
        return False
